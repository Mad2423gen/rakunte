import csv
import glob
import os
import sys
import random
import re
import shutil
import time
import lxml
import subprocess
import urllib.parse
from datetime import datetime
from functools import wraps
import requests
from requests.exceptions import Timeout
import schedule
# import win32com.client
from bs4 import BeautifulSoup
from joblib import Parallel, delayed
import add_functions
import logging
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment

# path definition============================================================
path = os.getcwd()
dat_dir = os.path.join(path, '../dat')  # csc,imgフォルダの親フォルダ
csv_dir = os.path.join(dat_dir, 'csv')  # csv保存フォルダ
img_dir = os.path.join(dat_dir, 'img')  # img保存フォルダ
output_dir = os.path.join(path, '../output')  # 出力フォルダ
conf_dir = os.path.join(path, '../config')  # テンプレート、タイムテーブルファイル保存フォルダ
keyword_dir = os.path.join(conf_dir, 'keyword')  # 除外キーワードディレクトリ
# エクセルテンプレートファイル
ex_temp_file = os.path.join(conf_dir, 'temp_file.xlsm')
# タイムテーブルファイル
time_table_file = os.path.join(conf_dir, 'timetable.csv')
# タイムスタンプファイル
time_stamp_file = os.path.join(conf_dir, 'time_tamp.txt')
# キーワードファイル　共通.txt
common_keyword_file = os.path.join(keyword_dir, '共通.txt')


# ===========================================================================
# 現在日時取得 get Current date and time
def add_datetime():
    return datetime.now().strftime('%Y-%m-%d-%H_%M_%S')


# ===========================================================================
# イメージファイル名生成　※不要になった
def filename_creation(src):
    return urllib.parse.urlparse(src)[2].replace('/', '')


# ===========================================================================
# image download　※不要になった
# When there is only one URL to link to
def img_save(src, save_dir=img_dir):
    os.makedirs(save_dir, exist_ok=True)
    while True:
        try:
            with open(os.path.join(save_dir, filename_creation(src)), "wb") as f:
                f.write(requests.get(src, timeout=(3, 3)).content)
            break
        except requests.exceptions.RequestException as e:
            print(e)
            print('***ダウンロードエラー リトライします***')
            print('*******Download error, retry*******')
            time.sleep(1)


# ===========================================================================
# If the linked URL is listed (parallel processing)
# ※複数のURLから画像をダウンロードするための並列処理　※不要になった
def img_saves(urls):
    # (withブロックを抜けるまで待機する)　n_jabs=n がスレッド数 verbose=1はメッセージ深度
    with Parallel(n_jobs=-1, verbose=1) as parallel:
        parallel(delayed(img_save)(url) for url in urls)


# ===========================================================================
# エクセルにエクスポート
# 引数： output_ex_files_dir == otput/time_interval_dir/datetime
def export_ex(output_ex_files_dir, intervaltime):

    # rakunte/output/intervaltime/日時フォルダに移動しての処理-------------------
    os.chdir(output_ex_files_dir)

    # for csv_file in csv_filenames:
    for csv_file in [os.path.basename(file_path)
                     for file_path in glob.glob(os.path.join(csv_dir, f'{intervaltime}_*.csv'))]:
        dt = add_functions.csv_read(os.path.join(csv_dir, csv_file))
        print(f'from {csv_file} export to excel')
        book = openpyxl.load_workbook(os.path.join(conf_dir, 'temp_file.xlsx'))
        sheet = book['Sheet1']
        for i, low in enumerate(dt):
            # img_link
            sheet.cell(row=i + 2, column=2).value = low[1]
            sheet.cell(row=i + 2, column=2).font = Font(color='FFFFFF')
            # title
            sheet.cell(row=i + 2, column=4).value = low[0]
            sheet.cell(row=i + 2, column=4).alignment = \
                Alignment(horizontal='justify', vertical='center')
            # title_link
            sheet.cell(row=i + 2, column=5).hyperlink = low[3]
            sheet.cell(row=i + 2, column=5).alignment = \
                Alignment(horizontal='justify', vertical='center')
            # price
            sheet.cell(row=i + 2, column=6).value = low[4]
            sheet.cell(row=i + 2, column=6).alignment = \
                Alignment(horizontal='justify', vertical='center')
            # review
            sheet.cell(row=i + 2, column=7).value = low[5]
            sheet.cell(row=i + 2, column=7).alignment = \
                Alignment(horizontal='justify', vertical='center')

        book.save(f'{os.path.splitext(csv_file)[0]}.xlsx')
        book.close()
        print('process termination')

        # # csvから画像ファイル名抽出、ファイル名抽出
        # print(f'{csv_file}をエクスポート')
        # # 作業用ファイルのダミーファイル
        # dummy_filename = fr'{random.randint(0, 100000)}.xlsm'
        # # フルパス
        # dummy_file_path = os.path.join(output_ex_files_dir, dummy_filename)
        #
        # # エクセルのテンプレートファイルをダミーファイル名でコピー（重複防止）
        # shutil.copyfile(ex_temp_file, dummy_file_path)
        #
        # # from csv to write exel
        # with open(os.path.join(csv_dir, csv_file),
        #           'r', encoding='utf-8_sig', newline='') as csvf:
        #     reader = csv.reader(csvf)
        #
        #
        # time.sleep(1)
        #
        # rename ダミーファイル名をジャンル名に
        # try:
        #     os.rename(dummy_filename, f'{os.path.splitext(csv_file)[0]}.xlsm')
        #     print('rename termination\n')
        # except:
        #     print('!!!With error rename-handling!!!')

    # excel spplication shutdown
    # excel.quit()

    # カレントディレクトリに復帰
    os.chdir(path)


# ===========================================================================
# Page source acquisition block
# スクレイピング本体、　１ページのソースからタグを検出、データ取得
# レビューとプライスを付加したバージョン
def scray_thumbnail(target_url):

    # time.sleep(0.25)
    res = requests.get(target_url, timeout=(30.0, 30.0))

    # javascriptとして扱われているhtmlコードを解除-->これで20位以降のソースも読み込めるようになる
    html_source = res.text.replace('<script language="JavaScript" type="text/javascript">', '')

    # parse
    soup = BeautifulSoup(html_source, 'lxml')

    # Confirmation of page existence
    flags = soup.find('img', src=re.compile('./指定されたページが見つかりません（エラー404）_ 楽天_files/w100.gif'))

    if flags:
        pass
    else:
        # Declaration of tag element list
        title_lists, title_urls, filenames, img_urls, revirews_lists, price_lists = [], [], [], [], [], []
        # Declaration of elements for output
        out_datas = []

        # hint review_tagはあったりなかったりするので、先ず親タグからその部分のブロックを抽出、
        # 更にfindないしselectで抽出する、返り値がFalseの場合はレビューが存在しないと言うこと

        # get title,title_url,review
        for title_block_source in soup.select('div.rnkRanking_upperbox'):

            # get litle,title_url
            title = title_block_source.select_one('div.rnkRanking_itemName > a')
            title_url = title.attrs['href']
            title_lists.append(str(title.text))
            title_urls.append(title_url)

            # get review  ※title_block内に「https://review.rakuten」が含まれなければNoneを返す
            if review_tag := title_block_source.find(href=re.compile('https://review.rakuten')):
                revirews_lists.append(review_tag.text.replace('レビュー(', '').replace('件)', ''))
            else:
                revirews_lists.append('None')

        # get pcrice
        [price_lists.append(p.text.replace('円', '')) for p in soup.select('div.rnkRanking_price')]
        # get img
        [img_urls.append(img_url.attrs['src']) for img_url
         in soup.select('div.rnkRanking_image > div > a > img')]

        #  Reference Element List============================================================
        #   entry ==  title, img_url, filename, title_url, price, review
        #
        #  for out_datas
        #  dataname == title_lists, title_urls, finames, img_urls, revirews_lists, price_list
        #  dataname == out_datas
        # ===================================================================================

        for i, title in enumerate(title_lists):
            # entry is --> title, img_url, filename, title_url, price, review
            # filenameのエントリーは不要になったので空欄、他のロジックに影響するので削除しないこと
            out_datas.append((title, img_urls[i], '', title_urls[i], price_lists[i], revirews_lists[i]))
            # for developer testing -------------
            # print(price_lists)
            # print(title_lists)
            # print(title_urls)
            # print(img_urls[1])
            # -----------------------------------
        return out_datas
        # return print(out_datas[0]) # for developer testing

# ===========================================================================
# スクレイピングとCSV保存、画像保存　※画像保存は無効化
def csv_save(genre, genre_id, intervaltime):
    # スクレイピング　（ジャンル内の全ページデータ取得）
    global old_csv_datas, keywords, exclusion_keywords, save_data
    print('\nrakuten scray')
    new_data = []
    for i in range(1, 5):  # 1~4ページ
        print(f'page:{i}')
        # target_url
        url = f'https://ranking.rakuten.co.jp/{intervaltime}/{genre_id}/p={i}'

        # スクレイピングしたデータを new_data に格納
        [new_data.append([ttl[0], ttl[1], ttl[2], ttl[3], ttl[4], ttl[5]]) for ttl in scray_thumbnail(url)]

    # =====================csv保存データ作成==================================

    # キーワードディレクトリ＆ファイル確認(消失していた場合は補完）
    add_functions.make_keyword_file_missing()

    # ジャンル別キーワードファイル
    exclusion_keyword_file = os.path.join(keyword_dir, f'{genre}.txt')
    exclusion_keyword_file2 = common_keyword_file

    # 古いジャンルcsvファイル無
    if not os.path.isfile(os.path.join(csv_dir, f'{intervaltime}_{genre}.csv')):
        print('ジャンルCSV無、', end='')
        # 除外キーワードファイル有
        if os.path.isfile(exclusion_keyword_file) and os.path.isfile(exclusion_keyword_file2):
            print('除外キーワードファイル有、', end='')
            keywords = add_functions.read_keywords(exclusion_keyword_file)
            # キーワード登録無
            if not keywords:
                # new_dataをそのまま保存
                print('キーワード登録無、新規データをジャンルcsvへ保存')
                save_data = new_data
            # キーワード登録有
            else:
                save_data = []
                print('キーワード登録有、該当を削除、ジャンルCSVへ保存')
                # new_dataからキーワード対象を除外する
                for ndata in new_data:
                    if not any(dta in ndata[0] for dta in keywords):
                        save_data.append(ndata)

        # 除外キーワードファイル無
        else:
            # new_dataをそのまま保存
            print('除外キーワードファイル無、新規データをそのまま保存')
            save_data = new_data

    # 古いジャンルcsvファイル有
    else:
        print('ジャンルCSV有、', end='')

        # 除外キーワードファイル無
        if not os.path.isfile(exclusion_keyword_file):
            # 古いジャンルcsvと新規データを比較、差分をnew_dataへ保存
            print('キーワードファイル無、新旧ジャンルcsvの差分をジャンルcsvへ保存')
            old_data = add_functions \
                .csv_read_title(os.path.join(csv_dir, f'{intervaltime}_{genre}.csv'))
            save_data = [nd for nd in new_data if not nd[0] in old_data]

        # 除外キーワードファイル有
        else:
            # キーワード登録無し
            keywords = add_functions.read_keywords(exclusion_keyword_file)
            print('除外キーワードファイル有、', end='')
            if not keywords:
                # 古いジャンルcsvと新規データを比較、差分をnew_dataへ保存
                print('キーワード登録無、新旧データを比較、差分を保存')
                old_data = add_functions \
                    .csv_read_title(os.path.join(csv_dir, f'{intervaltime}_{genre}.csv'))
                save_data = [nd for nd in new_data if not nd[0] in old_data]

            # キーワード登録有り
            else:
                # ジャンルcsvとキーワードを結合、new_dataと比較する
                print('キーワード登録有、登録キーワードを結合、新データと比較、差分を保存')
                save_data = []
                old_data = add_functions \
                    .csv_read_title(os.path.join(csv_dir, f'{intervaltime}_{genre}.csv'))
                joint_datas = old_data + keywords
                # 複数の条件のいずれにも当てはまらなければsave_dataに追加
                # save_data.extend([ndata for ndata in new_data
                #                   if not any(dta in ndata[0] for dta in joint_datas)])
                save_data = [ndata for ndata in new_data
                             if not any(dta in ndata[0] for dta in joint_datas)]

    # url重複判定（timetable.csv設定値が「URL_duplicate_detection,1」の場合ONになる
    save_data = add_functions.url_duplicate_detection(save_data, intervaltime, genre)

    # img&csv保存===========================================================

    # 画像の保存についてはエクセルのimage関数を使用して、セルにリンク先を貼れば画像を
    # 表示する機能が追加されていたので変更した。それにともなって画像処理部分は無効化した

    # 画像のダウンロード部　※不要な処理に追加
    # save_dataから画像のリンクを取得し、ダウンロード（並列処理）
    # print('img save now')
    # img_saves([row[1] for row in save_data])
    # print('img saved', '\n')

    # csvへ取得データ保存
    print('csv save now')
    with open(os.path.join(csv_dir, f'{intervaltime}_{genre}.csv'),
              'a', encoding='utf-8_sig', newline='') as sdf:
        csv.writer(sdf).writerows(save_data)
    print('csv saved\n')


# ===========================================================================
# 全処理実行関数　mode:リアルタイム、デイリー、ウィークリー選択　　mode2:テスト、本番実行選択
def main_func(mode=1, mode2=1):
    # error防止: 残っているエクセルタスクを強制終了
    print('\n excel task kill')
    subprocess.run('taskkill /F /T /IM excel.exe', stdout=None, shell=True)

    # 開始時刻取得
    start_time = add_datetime()
    print('\n==========スクレイピング処理開始==========')

    # counting period
    global intervaltime, genre_file
    intervaltime = 'realtime' if mode == 1 else 'daily' \
        if mode == 2 else 'weekly' if mode == 3 else 'unknown'

    print(f'スクレイピング範囲：{intervaltime}')
    # path definition----------------------------------------------
    # 取得ジャンル一覧を読み込み
    if mode2 == 1:  # テスト用
        print('\n=====debug mode=====\n')
        genre_file = os.path.join(path, r'../config/test_rakuten_genre.csv')
    elif mode2 == 2:  # 本実行
        print('\n=====main function=====\n')
        genre_file = os.path.join(path, r'../config/rakuten_genre.csv')
    # path definition----------------------------------------------

    # csvファイル更新日確認(datフォルダ消去、作成)
    print('Check update interval of csv files')

    # 新タイミングを読み込み、設定に従って削除する　
    # 設定日数取得
    specified_date = add_functions.csv_read(time_table_file)[3][1]

    # フォルダ作成、存在する場合はスキップ
    os.makedirs(csv_dir, exist_ok=True)
    os.makedirs(img_dir, exist_ok=True)
    # 作成日を参照し、期間が経過していたらcsv and img フォルダ削除
    add_functions.delete_old_files(specified_date)

    # 設定ファイルからジャンル、ジャンルID読み込み～スクレイピング～CSV作成
    for row in csv.reader(open(genre_file, 'r', encoding='utf-8_sig', newline='')):
        print(f'genre:{row[0]}   genre_id:{row[1]}')
        # スクレイピングとCSV書き込み
        csv_save(row[0], row[1], intervaltime)

    # エクセルへの書き込み実行=================================================
    # 【重要】マクロを実行して写真を貼り付ける場合、同じディレクトリにimgフォルダがないとエクセルの画像が
    #   消失してしまうので、画像ファイルを先にoutput+timespan+日付+imgフォルダに移動させ、
    #   テンプレートのエクセルファイルを移動させてからマクロを実行する。
    #   全部の処理後にリネームしないと、マクロがエラーを起こす

    # 処理後のエクセルファイルの保存先ディレクトリ　例）output/realtime/日付フォルダ
    output_ex_files_dir = os.path.join(output_dir, intervaltime, add_datetime())
    os.makedirs(output_ex_files_dir)

    # 写真ファイルのコピー dat/imgからoutpu/intervaltime/datetime/img
    # print('copy img files')
    # shutil.copytree(img_dir, os.path.join(output_ex_files_dir, 'img'))

    print('==========エクセル処理開始==========')
    # time.sleep(5)  # コピー終了待機
    # CSVからエクセルへ書き込み＆マクロ実行
    export_ex(output_ex_files_dir, intervaltime)

    # 処理時間result
    print('\n==========全処理終了==========')
    end_time = add_datetime()
    with open(time_stamp_file, 'a', encoding='utf-8_sig') as f:
        print(f'開始時間：{start_time}\n終了時間：{end_time}', file=f)
    print(f'開始時間：　{start_time}')
    print(f'終了時間：　{end_time}')


# ===========================================================================

if __name__ == '__main__':

    # 1はテスト、２は本番
    mode_b = 1

    # time_table import
    time_list = []
    [time_list.append(row)
     for row in csv.reader(open(time_table_file, 'r', encoding='utf-8_sig', newline=''))]

    if mode_b == 1:
        # ===for test===
        while True:
            main_func(mode=1, mode2=mode_b)
            print('\n*****待機中*****\n')
            time.sleep(600)

    elif mode_b == 2:
        # timer
        print(time_list)
        # realtime
        [schedule.every().day.at(t).do(main_func, mode=1, mode2=mode_b) for t in time_list[0][1:]]
        # daily
        [schedule.every().day.at(t2).do(main_func, mode=2, mode2=mode_b) for t2 in time_list[1][1:]]
        # weekly
        [schedule.every().monday.at(t3).do(main_func, mode=3, mode2=mode_b) for t3 in time_list[2][1:]]

        print('\n==== ratenk start====')
        print('==== 処理時間まで待機==')

        while True:
            schedule.run_pending()
            time.sleep(1)
    else:
        print('デバッグ指定エラー')
